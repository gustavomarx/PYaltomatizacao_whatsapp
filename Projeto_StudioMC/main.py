from PySimpleGUI import PySimpleGUI as sg
import pathlib
import openpyxl, xlrd
from openpyxl import Workbook

# Layout
sg.theme('DarkBlue4')

# Encontrar ou criar ficheiro

ficheiro = pathlib.Path(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")



if ficheiro.exists():
        pass
else:
    ficheiro = Workbook()
    sheets = ["Cadastro Atendimentos", "Estoque", "Controle Gastos", "Histórico Faturamento", "Histórico Comissões"]
    
    for s in sheets:
        ficheiro.create_sheet(s)

    ficheiro.move_sheet("Cadastro Atendimentos")
    
    folha=ficheiro.active
    ficheiro.remove_sheet(folha)
    ficheiro.save(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")

    #Criando sheet 'Cadastro Atendimentos'
    ficheiro.active = ficheiro['Cadastro Atendimentos']
    folha=ficheiro.active
    folha['A1']="Data atendimento"
    folha['B1']="Nome cliente"
    folha['C1']="Contato cliente"
    folha['D1']="Serviços realizados"
    folha['E1']="Total Comanda"
    folha['F1']="Agendou manutenção?"
    folha['G1']="Modelo?"
    folha['H1']="Profissional"
    folha['I1']="Forma de pagamento"

    ficheiro.save(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")

    #Criando sheet 'Estoque'
    ficheiro.active = ficheiro['Estoque']
    folha=ficheiro.active
    folha['A1']="Data entrada/saída"
    folha['B1']="Material"
    folha['C1']="Quantidade"
    folha['D1']="Profissional registrou"

    ficheiro.save(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")

    #Criando sheet 'Controle Gastos'
    ficheiro.active = ficheiro['Controle Gastos']
    folha=ficheiro.active
    folha['A1']="Data do gasto"
    folha['B1']="Descrição"
    folha['C1']="Valor em R$"
    folha['D1']="Forma pagamento"
    folha['E1']="Mês competência"
    folha['F1']="Categoria"

    ficheiro.save(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")

    #Criando sheet 'Histórico Faturamento'
    ficheiro.active = ficheiro['Histórico Faturamento']
    folha=ficheiro.active
    folha['A1']="Mês competência"
    folha['B1']="Faturamento total R$"
    folha['C1']="Número atendimentos"
    folha['D1']="Número clientes únicos"
    folha['E1']="Ticket médio R$"
    folha['F1']="Comissão paga R$"

    #Criando sheet 'Histórico Comissões'
    ficheiro.active = ficheiro['Histórico Comissões']
    folha=ficheiro.active
    folha['A1']="Mês competência"
    folha['B1']="Faturamento total R$"
    folha['C1']="Número atendimentos"
    folha['D1']="Comissão R$"
    folha['E1']="Ticket médio R$"
    folha['F1']="Bônus extra R$"


    ficheiro.save(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")




# INSERINDO DADOS TELA CADASTRO
def submit_cadastro(data_atendimento, nome_cliente, contato_cliente, servicos, total_comanda, profissional, forma_pagamento, agendou_manutencao, modelo):
    #pegando os dados dos entrys
            
    ficheiro = openpyxl.load_workbook(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")
    ficheiro.active = ficheiro['Cadastro Atendimentos']
    folha=ficheiro.active

    folha.cell(column=1, row=folha.max_row+1, value=data_atendimento)
    folha.cell(column=2, row=folha.max_row, value=nome_cliente)
    folha.cell(column=3, row=folha.max_row, value=contato_cliente)
    folha.cell(column=4, row=folha.max_row, value=servicos)
    folha.cell(column=5, row=folha.max_row, value=total_comanda)
    folha.cell(column=6, row=folha.max_row, value=agendou_manutencao)
    folha.cell(column=7, row=folha.max_row, value=modelo)
    folha.cell(column=8, row=folha.max_row, value=profissional)
    folha.cell(column=9, row=folha.max_row, value=forma_pagamento)

    ficheiro.save(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")

    sg.popup_auto_close('Dados salvos com sucesso!')

# INSERINDO DADOS TELA ESTOQUE
def submit_estoque(data, material, quantidade, profissional_registrou):
    #pegando os dados dos entrys
            
    ficheiro = openpyxl.load_workbook(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")
    ficheiro.active = ficheiro['Estoque']
    folha=ficheiro.active

    folha.cell(column=1, row=folha.max_row+1, value=data)
    folha.cell(column=2, row=folha.max_row, value=material)
    folha.cell(column=3, row=folha.max_row, value=quantidade)
    folha.cell(column=4, row=folha.max_row, value=profissional_registrou)

    ficheiro.save(r"cadastro_gastos_MCSJ\registros_projeto.xlsx")

    sg.popup_auto_close('Dados salvos com sucesso!')


# CRIANDO AS TELAS

def tela_login():  
    layout = [
        [sg.Text('SISTEMA DE GERENCIAMENTO MC SJ')],
        [sg.Text('Usuário'), sg.Push(), sg.Input(key='usuario', size=(20,1))],
        [sg.Text('Senha  '), sg.Push(), sg.Input(key='senha', password_char='*', size=(20,1))],
        [sg.Push(), sg.Button('Entrar'), sg.Push()]
    
    ]

    return sg.Window('Login', layout=layout, finalize=True)


def tela_menu():
    layout = [
        [sg.Push(), sg.Button('Cadastro atendimentos'), sg.Push()],
        [sg.Push(), sg.Button('Estoque'), sg.Push()],
        [sg.Push(), sg.Button('Controle gastos'), sg.Push()],
        [sg.Push(), sg.Button('Histórico faturamento'), sg.Push()],
        [sg.Push(), sg.Button('Histórico comissões'), sg.Push()],
   
    ]
    return sg.Window('Menu', layout=layout, finalize=True)


def tela_cadastro():
    profissionais = ["Anelise", "Brenda", "Bruna", "Eduarda", "Maria Luiza", "Yasmim", "Vanessa"]
    forma_pgto = ["Pix", "Dinheiro", "Cartão de Crédito", "Cartão de Débito"]
    
    layout = [
        [sg.Text('Data atendimento'), sg.Push(), sg.Input(key='data_atendimento', size=(30,1))],
        [sg.Text('Nome cliente'), sg.Push(), sg.Input(key='nome_cliente', size=(30,1))],
        [sg.Text('Contato cliente'), sg.Push(), sg.Input(key='contato_cliente', size=(30,1))],
        [sg.Text('Serviços realizados'), sg.Push(), sg.Input(key='servicos', size=(30,1))],
        [sg.Text('Total comanda'), sg.Push(), sg.Input(key='total_comanda', size=(30,1))],
        [sg.Text('Agendou manutenção?'), sg.Push(), sg.Radio('Sim',"agendou",key='agendou_manutencao_sim', enable_events=True, default=True), sg.Radio('Não',"agendou", key='agendou_manutencao_nao', enable_events=True)],
        [sg.Text('Modelo?'), sg.Push(), sg.Radio('Sim',"modelo",key='modelo_sim', enable_events=True, default=True), sg.Radio('Não',"modelo", key='modelo_nao', enable_events=True)],
        [sg.Text('Profissional'), sg.Push(), sg.Combo(profissionais, enable_events=True, readonly=True, default_value=profissionais[0], key="-profissionais-", size=(25,1))],
        [sg.Text('Forma de pagamento'), sg.Push(), sg.Combo(forma_pgto, enable_events=True, readonly=True, default_value=forma_pgto[0], key="-forma_pgto-", size=(25,1))],
        [sg.Push(), sg.Button('Cadastrar'), sg.Button('Voltar ao Menu'), sg.Push()]

    ]
    return sg.Window('Menu', layout=layout, finalize=True)


def tela_estoque():
    profissionais = ["Anelise", "Brenda", "Bruna", "Eduarda", "Maria Luiza", "Yasmim", "Vanessa"]
    layout = [
        [sg.Text('Data entrada/saída'), sg.Push(), sg.Input(key='data_entrada_saida', size=(30,1))],
        [sg.Text('Material'), sg.Push(), sg.Input(key='material', size=(30,1))],
        [sg.Text('Quantidade'), sg.Push(), sg.Input(key='quantidade', size=(30,1))],
        [sg.Text('Profissional registrou'), sg.Push(), sg.Combo(profissionais, enable_events=True, readonly=True, default_value=profissionais[0], key="-profissionais-", size=(25,1))],
        [sg.Push(), sg.Button('Entrada'), sg.Button('Saída'), sg.Push()],
        [sg.Push(), sg.Button('Voltar ao Menu'), sg.Push()]

    ]
    return sg.Window('Menu', layout=layout, finalize=True)


def tela_controle_gastos():
    mes = ["Jan", "Fev", "Mar", "Abr", "Maio", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    categoria = ["Alimentação", "Lazer", "Saúde", "Transporte", "Empresa", "Educação", "Investimento", "Outros"]
    forma_pgto = ["Pix", "Dinheiro", "Cartão Débito Nubank", "Cartão Crédito Nubank", "Cartão Crédito C6Bank", "Cartão Crédito PJ Nubank", "Cartão Crédito BB", "Boleto"]

    layout = [
        [sg.Text('Data do gasto'), sg.Push(), sg.Input(key='data_gasto', size=(30,1))],
        [sg.Text('Descrição'), sg.Push(), sg.Input(key='descricao', size=(30,1))],
        [sg.Text('Valor em R$'), sg.Push(), sg.Input(key='valor', size=(30,1))],
        [sg.Text('Forma pagamento'), sg.Push(), sg.Combo(forma_pgto, enable_events=True, readonly=True, default_value=forma_pgto[0], key="-profissionais-", size=(25,1))],
        [sg.Text('Mês competência'), sg.Push(), sg.Combo(mes, enable_events=True, readonly=True, default_value=mes[0], key="-profissionais-", size=(25,1))],
        [sg.Text('Categoria'), sg.Push(), sg.Combo(categoria, enable_events=True, readonly=True, default_value=categoria[0], key="-profissionais-", size=(25,1))],
        [sg.Push(), sg.Button('Cadastrar'), sg.Button('Voltar ao Menu'), sg.Push()]

    ]
    return sg.Window('Menu', layout=layout, finalize=True)


def tela_historico_faturamento_studio():
    mes = ["Jan", "Fev", "Mar", "Abr", "Maio", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

    layout = [
        [sg.Text('Mês competência'), sg.Push(), sg.Combo(mes, enable_events=True, readonly=True, default_value=mes[0], key="-profissionais-", size=(25,1))],
        [sg.Text('Faturamento total R$'), sg.Push(), sg.Input(key='faturamento_total', size=(30,1))],
        [sg.Text('Número atendimentos'), sg.Push(), sg.Input(key='numero_atendimentos', size=(30,1))],
        [sg.Text('Número clientes únicos'), sg.Push(), sg.Input(key='numero_clientes_unicos', size=(30,1))],
        [sg.Text('Ticket médio R$'), sg.Push(), sg.Input(key='ticket_medio', size=(30,1))],
        [sg.Text('Comissão paga R$'), sg.Push(), sg.Input(key='comissao_paga', size=(30,1))],
        [sg.Push(), sg.Button('Cadastrar'), sg.Button('Voltar ao Menu'), sg.Push()]

    ]
    return sg.Window('Menu', layout=layout, finalize=True)

def tela_historico_faturamento_profissional():
    layout = [
        [sg.Text('Mês competência'), sg.Push(), sg.Input(key='mes_competencia', size=(30,1))],
        [sg.Text('Faturamento total R$'), sg.Push(), sg.Input(key='faturamento_total', size=(30,1))],
        [sg.Text('Número atendimentos'), sg.Push(), sg.Input(key='numero_atendimentos', size=(30,1))],
        [sg.Text('Comissão R$'), sg.Push(), sg.Input(key='numero_clientes_unicos', size=(30,1))],
        [sg.Text('Ticket médio R$'), sg.Push(), sg.Input(key='ticket_medio', size=(30,1))],
        [sg.Text('Bônus extra R$'), sg.Push(), sg.Input(key='comissao_paga', size=(30,1))],
        [sg.Push(), sg.Button('Cadastrar'), sg.Button('Voltar ao Menu'), sg.Push()]

    ]
    return sg.Window('Menu', layout=layout, finalize=True)





# Criando as janelas iniciais

tela1, tela2, tela3, tela4, tela5, tela6, tela7 = tela_login(), None, None, None, None, None, None

# Ler os eventos

while True:
    window, eventos, valores = sg.read_all_windows()

    # Quando a janela for fechada
    if window == tela1 and eventos == sg.WIN_CLOSED:
        break

    if window == tela2 and eventos == sg.WIN_CLOSED:
        break

    if window == tela3 and eventos == sg.WIN_CLOSED:
        break

    if window == tela4 and eventos == sg.WIN_CLOSED:
        break

    if window == tela5 and eventos == sg.WIN_CLOSED:
        break

    if window == tela6 and eventos == sg.WIN_CLOSED:
        break

    if window == tela7 and eventos == sg.WIN_CLOSED:
        break

    # LOGIN
    if window == tela1 and eventos == 'Entrar':
        if valores['usuario'] == 'admin' and valores['senha'] == '123':
            print("Bem vindo")
            tela2 = tela_menu()
            tela1.hide()
        else: 
            print("Senha errada")


    # ESCOLHA DO MENU
    if window == tela2 and eventos == 'Cadastro atendimentos':
        print("entrou cadastro")
        tela3 = tela_cadastro()
        tela2.hide()
        

    if window == tela2 and eventos == 'Estoque':
        print("entrou estoque")
        tela4 = tela_estoque()
        tela2.hide()

    if window == tela2 and eventos == 'Controle gastos':
        print("entrou gasto")
        tela5 = tela_controle_gastos()
        tela2.hide()

    if window == tela2 and eventos == 'Histórico faturamento':
        print("entrou faturmaneot")
        tela6 = tela_historico_faturamento_studio()
        tela2.hide()

    if window == tela2 and eventos == 'Histórico comissões':
        print("entrou faturmaneto2")
        tela7 = tela_historico_faturamento_profissional()
        tela2.hide()

    # VOLTAR AO MENU

    if window == tela3 and eventos == 'Voltar ao Menu':
        tela3.hide()
        tela2.un_hide()

    if window == tela4 and eventos == 'Voltar ao Menu':
        tela4.hide()
        tela2.un_hide()

    if window == tela5 and eventos == 'Voltar ao Menu':
        tela5.hide()
        tela2.un_hide()

    if window == tela6 and eventos == 'Voltar ao Menu':
        tela6.hide()
        tela2.un_hide()

    if window == tela7 and eventos == 'Voltar ao Menu':
        tela7.hide()
        tela2.un_hide()


    # PEGANDO DADOS TELA CADASTRO
    if window == tela3 and eventos == 'Cadastrar':

        data_atendimento = valores['data_atendimento']
        nome_cliente = valores['nome_cliente']
        contato_cliente = valores['contato_cliente']
        servicos = valores['servicos']
        total_comanda = valores['total_comanda']
        profissional = valores['-profissionais-']
        forma_pagamento = valores['-forma_pgto-']
        
        if(valores['agendou_manutencao_sim'] == True):
            agendou_manutencao = 'sim'
        elif(valores['agendou_manutencao_nao'] == True):
            agendou_manutencao = 'nao'

        if(valores['modelo_sim'] == True):
            modelo = 'sim'
        elif(valores['modelo_nao'] == True):
            modelo = 'nao'

        if(data_atendimento == '' or nome_cliente == '' or contato_cliente == '' or servicos == '' or total_comanda == '' or profissional == '' or forma_pagamento == ''):
            sg.popup_auto_close('Informe todos os campos!')

        else:

            submit_cadastro(data_atendimento, nome_cliente, contato_cliente, servicos, total_comanda, profissional, forma_pagamento, agendou_manutencao, modelo)

            #limpando campos
            window['nome_cliente'].update(valores['nome_cliente'][:-50])
            window['contato_cliente'].update(valores['contato_cliente'][:-50])
            window['servicos'].update(valores['servicos'][:-50])
            window['total_comanda'].update(valores['total_comanda'][:-50])

            print(data_atendimento)
            print(nome_cliente)
            print(contato_cliente)
            print(servicos)
            print(total_comanda)
            print(profissional)
            print(forma_pagamento)
            print(agendou_manutencao)
            print(modelo)


    # TELA ESTOQUE
    if window == tela4 and eventos == 'Entrada':

        data = valores['data_entrada_saida']
        material = valores['material']
        quantidade = valores['quantidade']
        profissional_registrou = valores['-profissionais-']

        if(data == '' or material == '' or quantidade == '' or profissional_registrou == ''):
            sg.popup_auto_close('Informe todos os campos!')

        else:
            submit_estoque(data, material, quantidade, profissional_registrou)

            #limpando campos
            window['material'].update(valores['material'][:-50])
            window['quantidade'].update(valores['quantidade'][:-50])

            print(data)
            print(material)
            print(quantidade)
            print(profissional_registrou)







        

        


            