import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

# Aparência padrão do sistema

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.todo_sistema()

    def layout_config(self):
        self.title("Sistema de cadastro de gastos Meus Cílios - São José")
        self.geometry("700x700")

    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', '#fff']).place(x=50,
                                                                                                                 y=640)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50,
                                                                                                                  y=660)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal").place(x=0,
                                                                                                                  y=0)
        title = ctk.CTkLabel(frame, text="Sistema de gastos MC São José", font=("Ubuntu bold", 24),
                             fg_color="transparent", bg_color="teal", text_color="#fff")
        title.place(x=190, y=10)

        span = ctk.CTkLabel(frame, text="Por favor, preencha todos os campos do formulário", font=("Ubuntu bold", 16),
                            text_color=["#000", "#fff"]).place(x=50, y=70)

        ficheiro = pathlib.Path(r"C:\Users\gusta\OneDrive\Documentos\Relatorio de gastos\relatorio_gastos.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = "Data"
            folha['B1'] = "Descrição"
            folha['C1'] = "Valor"
            folha['D1'] = "Forma Pagamento"
            folha['E1'] = "Mês Competência"
            folha['F1'] = "Categoria"

            ficheiro.save(r"C:\Users\gusta\OneDrive\Documentos\Relatorio de gastos\relatorio_gastos.xlsx")

        def submit():

            # pegando os dados dos entrys
            data = data_value.get()
            descricao = descricao_value.get()
            valor = valor_value.get()
            forma_pagamento = forma_pagamento_combobox.get()
            mes_competencia = mes_competencia_combobox.get()
            categoria = categoria_combobox.get()

            if (data == "" or descricao == "" or valor == "" or mes_competencia == ""):
                messagebox.showerror("Sistema", "Campos vazios")
            else:
                ficheiro = openpyxl.load_workbook(r"C:\Users\gusta\OneDrive\Documentos\Relatorio de gastos\relatorio_gastos.xlsx")
                folha = ficheiro.active

                folha.cell(column=1, row=folha.max_row + 1, value=data)
                folha.cell(column=2, row=folha.max_row, value=descricao)
                folha.cell(column=3, row=folha.max_row, value=valor)
                folha.cell(column=4, row=folha.max_row, value=forma_pagamento)
                folha.cell(column=5, row=folha.max_row, value=mes_competencia)
                folha.cell(column=6, row=folha.max_row, value=categoria)

                ficheiro.save(r"C:\Users\gusta\OneDrive\Documentos\Relatorio de gastos\relatorio_gastos.xlsx")

                messagebox.showinfo("Sistema", "Dados salvos com sucesso!")
                clear()

        def clear():

            descricao = descricao_value.set("")
            valor = valor_value.set("")

        # Variáveis
        data_value = StringVar()
        descricao_value = StringVar()
        valor_value = StringVar()

        # Entry
        data_entry = ctk.CTkEntry(self, width=100, font=("Ubuntu bold", 16), textvariable=data_value,
                                  fg_color="transparent")
        data_entry.insert(0, "dd/mm/yyyy")

        descricao_entry = ctk.CTkEntry(self, width=350, font=("Ubuntu bold", 16), textvariable=descricao_value,
                                       fg_color="transparent")
        valor_entry = ctk.CTkEntry(self, width=70, font=("Ubuntu bold", 16), textvariable=valor_value,
                                   fg_color="transparent")

        # Combobox
        forma_pagamento_combobox = ctk.CTkComboBox(self, values=["Pix", "Dinheiro", "Cartão Débito Nubank",
                                                                 "Cartão Crédito Nubank", "Cartão Crédito C6Bank",
                                                                 "Cartão Crédito PJ Nubank", "Cartão Crédito BB",
                                                                 "Boleto"], font=("Ubuntu bold", 14))
        forma_pagamento_combobox.set("Pix")

        mes_competencia_combobox = ctk.CTkComboBox(self,
                                                   values=["Jan", "Fev", "Mar", "Abr", "Maio", "Jun", "Jul", "Ago",
                                                           "Set", "Out", "Nov", "Dez"], font=("Ubuntu bold", 14))
        mes_competencia_combobox.set("Jan")

        categoria_combobox = ctk.CTkComboBox(self, values=["Alimentação", "Lazer", "Saúde", "Transporte", "Empresa",
                                                           "Educação", "Investimento", "Outros"],
                                             font=("Ubuntu bold", 14))
        categoria_combobox.set("Empresa")

        # Calendar

        # Labels
        lb_data = ctk.CTkLabel(frame, text="Informe a data", font=("Ubuntu bold", 16), text_color=["#000", "#fff"])
        lb_descricao = ctk.CTkLabel(frame, text="Descrição do gasto", font=("Ubuntu bold", 16),
                                    text_color=["#000", "#fff"])
        lb_valor = ctk.CTkLabel(frame, text="Valor em R$", font=("Ubuntu bold", 16), text_color=["#000", "#fff"])
        lb_forma_pagamento = ctk.CTkLabel(frame, text="Forma de pagamento", font=("Ubuntu bold", 16),
                                          text_color=["#000", "#fff"])
        lb_mes_competencia = ctk.CTkLabel(frame, text="Mês Competência", font=("Ubuntu bold", 16),
                                          text_color=["#000", "#fff"])
        lb_categoria = ctk.CTkLabel(frame, text="Categoria", font=("Ubuntu bold", 16), text_color=["#000", "#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar".upper(), command=submit, fg_color="#151",
                                   hover_color="#131").place(x=50, y=600)
        btn_clear = ctk.CTkButton(self, text="Limpar".upper(), command=clear, fg_color="#555",
                                  hover_color="#333").place(x=200, y=600)

        # Posicao dos elementos na janela
        lb_data.place(x=50, y=120)
        data_entry.place(x=50, y=160)

        lb_descricao.place(x=50, y=200)
        descricao_entry.place(x=50, y=230)

        lb_valor.place(x=50, y=270)
        valor_entry.place(x=50, y=300)

        lb_forma_pagamento.place(x=50, y=340)
        forma_pagamento_combobox.place(x=50, y=370)

        lb_mes_competencia.place(x=50, y=410)
        mes_competencia_combobox.place(x=50, y=440)

        lb_categoria.place(x=50, y=480)
        categoria_combobox.place(x=50, y=510)

    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)


if __name__ == "__main__":
    app = App()
    app.mainloop()